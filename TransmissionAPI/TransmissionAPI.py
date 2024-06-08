import transmission_rpc as rpc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json

class TransmissionParser:
    def __init__(self, host: str = 'localhost', port: int = 9091, username: str = '', pwd: str = '') -> None:
        self.host = host
        self.port = port
        self.username = username
        self.pwd = pwd
        self.torrents = []
        self.result = {}
        self.ok = False

    def get_torrents(self) -> None:
        try:
            self.result = {}
            with rpc.Client(host=self.host, port=self.port, username=self.username, password=self.pwd) as client:
                gt = client.get_torrents()
            for torrent in gt:
                size, size_type = self.human_readable_size(torrent.size_when_done)
                temp = {torrent.hashString[-6:]: {'hash': torrent.hashString, 'name': torrent.name,
                                                  'size': '{:.2f}'.format(size) + ' ' + size_type,
                                                  'trackers': torrent.tracker_list,
                                                  'torrent': torrent.torrent_file.replace('\\', '/'),
                                                  'download_dir': torrent.download_dir}}
                self.result.update(temp)
                self.ok = True
        except Exception as e:
            print(e)

    def to_json(self) -> str:
        if not self.ok:
            self.get_torrents()
        with open('transmission_torrent.json', 'w', encoding='utf-8')as f:
            f.write(json.dumps(self.result, indent=4, ensure_ascii=False))
        return json.dumps(self.result, indent=4, ensure_ascii=False)

    def to_excel(self):
        if not self.ok:
            self.get_torrents()
        temp = [['tid', 'name', 'size', 'trackers', 'hash', 'download_dir']]
        for i in self.result.keys():
            trackers = str(self.result[i]["trackers"]).replace("'", "").replace("[", "").replace("]", "")
            temp.append([i, self.result[i]["name"], self.result[i]["size"], trackers, 
            self.result[i]["hash"], self.result[i]["download_dir"]])
        wb = Workbook()
        ws = wb.active
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for row in temp:
            ws.append(row)
        for cell in ws['1']:
            cell.font = header_font
            cell.fill = header_fill
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        wb.save('transmission_torrent.xlsx')

    @staticmethod
    def human_readable_size(size_in_bytes: int):
        size_type = ['iB', 'KiB', 'MiB', 'GiB', 'TiB', 'PiB']
        for i, stype in enumerate(size_type):
            if size_in_bytes < 1024 or i == len(size_type) - 1:
                return size_in_bytes, stype
            size_in_bytes /= 1024

